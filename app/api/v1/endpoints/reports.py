from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import datetime
from typing import Optional
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from app.core.database import get_db
from app.core.security import get_current_active_user
from app.models.user import User
from app.models.sample import Sample

router = APIRouter()

@router.get("/samples")
async def export_samples_report(
    format: str = Query("excel", regex="^(excel|pdf)$"),
    sample_type: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Export samples to Excel or PDF"""
    # Only admins and supervisors can export
    if current_user.role not in ["admin", "supervisor"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Build query
    from sqlalchemy import select, and_
    query = select(Sample)
    
    filters = []
    if sample_type:
        filters.append(Sample.sample_type == sample_type)
    if status:
        filters.append(Sample.status == status)
    if date_from:
        filters.append(Sample.collected_at >= date_from)
    if date_to:
        filters.append(Sample.collected_at <= date_to)
    
    if filters:
        query = query.where(and_(*filters))
    
    result = await db.execute(query)
    samples = result.scalars().all()
    
    if not samples:
        raise HTTPException(status_code=404, detail="No samples found")
    
    # Convert to dict
    data = []
    for s in samples:
        data.append({
            "sample_id": s.sample_id,
            "sample_type": s.sample_type.replace("_", " ").title(),
            "status": s.status.replace("_", " ").title(),
            "collected_at": s.collected_at.strftime("%Y-%m-%d %H:%M") if s.collected_at else "",
            "collection_point": s.collection_point or "",
            "is_anomaly": "⚠️ Yes" if s.is_anomaly else "✓ No",
            "result_data": str(s.result_data) if s.result_data else "",
        })
    
    if format == "excel":
        return await export_excel(data)
    else:
        return await export_pdf(data)

async def export_excel(data):
    """Generate Excel file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Samples Report"
    
    headers = ["Sample ID", "Type", "Status", "Collected At", "Collection Point", "Anomaly", "Results"]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for row_idx, item in enumerate(data, 2):
        row = [
            item["sample_id"],
            item["sample_type"],
            item["status"],
            item["collected_at"],
            item["collection_point"],
            item["is_anomaly"],
            item["result_data"],
        ]
        for col, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=samples_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

async def export_pdf(data):
    """Generate PDF file"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=20)
    
    elements = []
    elements.append(Paragraph("Samples Report", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # Table headers
    table_data = [["Sample ID", "Type", "Status", "Collected", "Anomaly"]]
    
    for item in data[:20]:  # Limit to 20 rows for PDF
        table_data.append([
            item["sample_id"],
            item["sample_type"],
            item["status"],
            item["collected_at"][:10],
            item["is_anomaly"],
        ])
    
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=samples_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )